import csv
import json
import os
import re
import threading
import time
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional

from logging_config import get_logger

logger = get_logger(__name__)

data_dir = os.path.join(os.path.dirname(__file__), "data")
if not os.path.exists(data_dir):
    os.makedirs(data_dir)

messages_dir = os.path.join(data_dir, "messages")
if not os.path.exists(messages_dir):
    os.makedirs(messages_dir)

group_manage_config_file = os.path.join(data_dir, "group_manage.json")

collection_config_file = os.path.join(data_dir, "collection_config.json")

monitoring_config_file = os.path.join(data_dir, "monitoring_config.json")

RECORDING_DIR = os.path.join(data_dir, "recordings")
if not os.path.exists(RECORDING_DIR):
    os.makedirs(RECORDING_DIR)

CHAT_DATE_DIR = os.path.join(os.path.dirname(__file__), "chat_date")
if not os.path.exists(CHAT_DATE_DIR):
    os.makedirs(CHAT_DATE_DIR)

DATA_DIR = "data"
MESSAGE_TYPES = ["friend", "group"]
SYSTEM_MESSAGE_TYPE = "sys"
SELF_SENDER = "Self"


def get_daily_messages_file(chat_name: str) -> str:
    date_str = datetime.now().strftime("%Y-%m-%d")
    safe_chat_name = re.sub(r'[\\/:*?"<>|]', '_', chat_name)
    filename = f"{safe_chat_name}_{date_str}.csv"
    return os.path.join(CHAT_DATE_DIR, filename)


class MessageInfo:
    def __init__(self, message: Any):
        self.message = message
    
    def get_sender(self) -> str:
        if hasattr(self.message, 'sender'):
            return self.message.sender
        elif hasattr(self.message, 'get') and callable(getattr(self.message, 'get')):
            return self.message.get('sender', '')
        return ""
    
    def get_content(self) -> str:
        if hasattr(self.message, 'content'):
            return self.message.content or ""
        elif hasattr(self.message, 'get') and callable(getattr(self.message, 'get')):
            return self.message.get("content", "")
        return str(self.message)
    
    def get_type(self) -> str:
        if hasattr(self.message, "type"):
            return self.message.type.lower()
        return ""
    
    def is_system_message(self) -> bool:
        return self.get_type() == SYSTEM_MESSAGE_TYPE
    
    def is_self_message(self) -> bool:
        return self.get_sender() == SELF_SENDER
    
    def is_empty_content(self) -> bool:
        return not self.get_content().strip()
    
    def is_valid_message_type(self) -> bool:
        return self.get_type() in MESSAGE_TYPES


class WorkerConfig:
    
    def __init__(
        self,
        wx_instance: Any,
        receiver: str,
        record_interval: int = 0,
        sensitive_words: List[str] = None
    ):
        self.wx_instance = wx_instance
        self.receiver = receiver
        self.record_interval = record_interval
        self.sensitive_words = sensitive_words or []


class MessageHistory:
    
    def __init__(self, sender: str, message: str, chat_name: str):
        self.sender = sender
        self.message = message
        self.chat_name = chat_name
        self.timestamp = datetime.now().isoformat()
        self.id = str(uuid.uuid4())
        self.time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            "sender": self.sender,
            "message": self.message,
            "chatName": self.chat_name,
            "timestamp": self.timestamp,
            "id": self.id,
            "time": self.time
        }


class WorkerState:
    
    def __init__(self):
        self.listen_list: List[str] = []
        self.last_record_info = {"content": "", "time": 0}
        self._message_lock = threading.Lock()
        self._stop_event = threading.Event()
        self._is_running = False
        self._paused = False
        self._pause_cond = threading.Condition(threading.Lock())
        self.start_time: Optional[float] = None
    
    def get_message_lock(self) -> threading.Lock:
        return self._message_lock
    
    def get_stop_event(self) -> threading.Event:
        return self._stop_event
    
    def is_running(self) -> bool:
        return self._is_running
    
    def set_running(self, running: bool) -> None:
        self._is_running = running
    
    def is_paused(self) -> bool:
        return self._paused
    
    def set_paused(self, paused: bool) -> None:
        self._paused = paused
    
    def get_pause_condition(self) -> threading.Condition:
        return self._pause_cond
    
    def get_start_time(self) -> Optional[float]:
        return self.start_time
    
    def set_start_time(self, start_time: Optional[float]) -> None:
        self.start_time = start_time


class GroupWorkerThread:
    def __init__(self, config: WorkerConfig):
        """
        初始化群聊管理工作线程

        Args:
            config: 工作线程配置对象

        Raises:
            ValueError: 当wx_instance参数为空时抛出
        """
        if not config.wx_instance:
            raise ValueError("wx_instance参数不能为空")

        self.config = config
        self.at_me = f"@{self.config.wx_instance.nickname}"
        self.receiver_list = [
            r.strip() for r in self.config.receiver.replace("，", ",").split(",") if r.strip()
        ]

        if os.path.exists(group_manage_config_file):
            with open(group_manage_config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)

            config["group"] = self.config.receiver

            with open(group_manage_config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)

        
        # 初始化状态对象
        self.state = WorkerState()
        
        # 初始化敏感词列表
        self.sensitive_words = self.config.sensitive_words or []
        
        # 初始化正则规则列表
        self.regex_rules = self._load_regex_rules()

        logger.info(
            "Group worker thread initialized: receiver=%s, record_interval=%ss",
            self.config.receiver,
            self.config.record_interval,
        )

    def update_rules(self) -> bool:
        """更新规则（空实现，因为不需要规则处理）
        
        Returns:
            bool: 始终返回False
        """
        return False

    def init_listeners(self) -> bool:
        """
        初始化消息监听器

        Returns:
            bool: True表示初始化成功，False表示失败
        """
        if self._should_stop():
            logger.warning("Listener initialization interrupted")
            return False

        for target in self.receiver_list:
            if self._should_stop():
                return False

            try:
                self.config.wx_instance.AddListenChat(who=target)
                self.state.listen_list.append(target)
                logger.info("Added listener: %s", target)
            except (ValueError, TypeError, AttributeError, KeyError) as e:
                logger.error("Failed to add listener: %s, error: %s", target, str(e))
                # 清理已添加的监听器
                self._cleanup()
                return False
            except OSError as e:
                logger.error("OS error adding listener: %s, error: %s", target, str(e))
                # 清理已添加的监听器
                self._cleanup()
                return False
            except RuntimeError as e:
                logger.error("Runtime error adding listener: %s, error: %s", target, str(e))
                # 清理已添加的监听器
                self._cleanup()
                return False
        return True

    def _get_chat_name(self, who: str) -> str:
        """
        获取聊天名称

        Args:
            who: 聊天对象标识

        Returns:
            str: 聊天名称，如果无法获取则返回原标识
        """
        if not hasattr(self.config.wx_instance, "GetChatName"):
            return who
        return self.config.wx_instance.GetChatName(who)

    def _is_target_message(self, message: Any) -> bool:
        """
        检查消息是否来自目标联系人

        Args:
            message: 消息对象

        Returns:
            bool: True表示来自目标联系人，False表示不是
        """
        try:
            msg_info = MessageInfo(message)
            sender = msg_info.get_sender()
            # 检查发送者是否在接收者列表中
            return sender in self.receiver_list
        except (ValueError, TypeError, AttributeError, KeyError) as e:
            logger.error("Error checking target message: %s", e)
            return False
        except OSError as e:
            logger.error("OS error checking target message: %s", e)
            return False
        except Exception as e:
            logger.error("Unexpected error checking target message: %s", e)
            return False



    def _append_to_daily_file(self, content: str, chat_name: str) -> None:
        """
        将内容追加到按天存储的消息文件（CSV格式）
        
        Args:
            content: 要追加的内容（格式：[时间] 发送者: {sender} | 群聊: {chat_name} | 内容: {message_content}）
            chat_name: 聊天名称
        """
        try:
            daily_file = get_daily_messages_file(chat_name)
            
            # 解析消息内容
            # 格式示例：[2025-09-21 10:30:25] 发送者: 张三 | 群聊: 测试群 | 内容: 你好
            import re
            pattern = r'\[(.*?)\] 发送者: (.*?) \| 群聊: (.*?) \| 内容: (.*)'
            match = re.match(pattern, content)
            
            if match:
                time_str, sender, chat_type, message_content = match.groups()
                
                # 准备CSV行数据
                row_data = {
                    '时间': time_str,
                    '发送者': sender,
                    '聊天类型': chat_type,
                    '聊天名称': chat_name,
                    '消息内容': message_content
                }
                
                # 检查文件是否存在
                file_exists = os.path.exists(daily_file)
                
                # 写入CSV文件
                with open(daily_file, 'a', encoding='utf-8', newline='') as f:
                    writer = csv.DictWriter(f, fieldnames=['时间', '发送者', '聊天类型', '聊天名称', '消息内容'])
                    
                    # 如果文件不存在，写入表头
                    if not file_exists:
                        writer.writeheader()
                        logger.info(f"创建新的按天存储CSV文件: {daily_file}")
                    
                    # 写入数据行
                    writer.writerow(row_data)
            else:
                # 如果格式不匹配，使用原始内容作为消息内容
                row_data = {
                    '时间': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    '发送者': '未知',
                    '聊天类型': '未知',
                    '聊天名称': chat_name,
                    '消息内容': content
                }
                
                # 检查文件是否存在
                file_exists = os.path.exists(daily_file)
                
                # 写入CSV文件
                with open(daily_file, 'a', encoding='utf-8', newline='') as f:
                    writer = csv.DictWriter(f, fieldnames=['时间', '发送者', '聊天类型', '聊天名称', '消息内容'])
                    
                    # 如果文件不存在，写入表头
                    if not file_exists:
                        writer.writeheader()
                        logger.info(f"创建新的按天存储CSV文件: {daily_file}")
                    
                    # 写入数据行
                    writer.writerow(row_data)
                    
        except IOError as e:
            logger.error(f"写入按天存储CSV文件失败: {e}")
        except Exception as e:
            logger.error(f"处理消息内容失败: {e}")

    def _check_sensitive_words(self, content: str) -> bool:
        """
        检查内容是否包含敏感词
        
        Args:
            content: 要检查的内容
            
        Returns:
            bool: 是否包含敏感词
        """
        if not self.sensitive_words:
            return False
            
        # 检查是否包含任何敏感词
        for word in self.sensitive_words:
            if re.search(re.escape(word), content, re.IGNORECASE):
                return True
        return False

    def _load_regex_rules(self) -> List[Dict[str, str]]:
        """加载正则规则列表
        
        Returns:
            List[Dict[str, str]]: 正则规则列表
        """
        try:
            if os.path.exists(regex_config_file):
                with open(regex_config_file, 'r', encoding='utf-8') as f:
                    config_data = json.load(f)
                    # 提取rules数组
                    return config_data.get("rules", [])
            return []
        except Exception as e:
            logger.error(f"加载正则规则失败: {e}")
            return []

    def _match_regex_rules(self, message_content: str) -> List[Dict[str, Any]]:
        """使用正则规则匹配消息内容
        
        Args:
            message_content: 消息内容
            
        Returns:
            List[Dict[str, Any]]: 匹配到的规则和提取的内容列表
        """
        matched_rules = []
        
        for rule in self.regex_rules:
            pattern = rule.get("pattern", "")
            if not pattern:
                continue
                
            try:
                # 使用正则表达式匹配
                match = re.search(pattern, message_content)
                if match:
                    # 提取所有匹配组的内容
                    extracted_contents = []
                    if len(match.groups()) > 0:
                        # 如果有捕获组，提取所有组的内容
                        extracted_contents = list(match.groups())
                    else:
                        # 如果没有捕获组，使用完整匹配
                        extracted_contents = [match.group(0)]
                    
                    # 将提取的内容用逗号分隔
                    extracted_content_str = ", ".join(extracted_contents)
                    
                    matched_rules.append({
                        "pattern": pattern,
                        "original_message": rule.get("originalMessage", ""),
                        "extracted_content": extracted_content_str,
                        "extracted_contents": extracted_contents,  # 保存原始提取内容列表
                        "full_match": match.group(0)
                    })
            except re.error as e:
                logger.error(f"正则表达式错误 '{pattern}': {e}")
            except Exception as e:
                logger.error(f"正则匹配时发生错误: {e}")
        
        return matched_rules

    def _save_matched_content(self, matched_data: Dict[str, Any], chat_name: str) -> None:
        """保存匹配到的内容到collect目录（CSV格式）
        
        Args:
            matched_data: 匹配到的数据
            chat_name: 聊天名称
        """
        try:
            # 创建collect目录
            collect_dir = os.path.join(CHAT_DATE_DIR, "collect")
            if not os.path.exists(collect_dir):
                os.makedirs(collect_dir)
            
            # 生成文件名：群聊名_YYYY-MM-DD.csv
            date_str = datetime.now().strftime("%Y-%m-%d")
            filename = f"{chat_name}_{date_str}.csv"
            filepath = os.path.join(collect_dir, filename)
            
            # 检查文件是否存在，如果不存在则写入表头
            file_exists = os.path.exists(filepath)
            
            # 获取提取内容列表
            extracted_contents = matched_data.get('extracted_contents', [])
            
            # 将提取内容列表转换为字符串表示，例如：[项目1, 项目2, 项目3]
            extracted_content_str = str(extracted_contents)
            
            # 准备CSV行数据
            row_data = {
                '时间': matched_data['time'],
                '发送者': matched_data['sender'],
                '群聊': chat_name,
                '原始消息': matched_data['full_match'],
                '匹配规则': matched_data['pattern'],
                '提取内容': extracted_content_str,
            }
            
            # 固定表头
            fieldnames = ['时间', '发送者', '群聊', '原始消息', '匹配规则', '提取内容']
            
            # 写入CSV文件
            with open(filepath, 'a', encoding='utf-8', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)                
                # 写入数据行
                writer.writerow(row_data)
                
            logger.info(f"[正则匹配] 已保存匹配内容到CSV文件: {filepath}")
            
        except Exception as e:
            print(e)
            logger.error(f"保存匹配内容到CSV失败: {e}")

    def _process_message(self, msg: Any = None, chat: Any = None) -> None:
        """
        Process received messages, only responsible for recording, no replies
        
        Args:
            msg: Message content
            chat: Message window
        """
        try:
            if msg is None:
                return
                
            chat_info = self._get_chat_info(chat)
            is_group = chat_info["type"] == "group"
            chat_name = chat_info["name"] or "Unknown chat"

            receive_time = time.time()
            time_str = datetime.fromtimestamp(receive_time).strftime("%Y-%m-%d %H:%M:%S")

            msg_info = MessageInfo(msg)
            message_content = msg_info.get_content()
            sender = msg_info.get_sender()

            if self._check_sensitive_words(message_content):
                logger.warning("Sensitive content detected: %s... Sender: %s", message_content[:50], sender)

            message_entry = f"[{time_str}] Sender: {sender} | Group: {chat_name if is_group else 'Private chat'} | Content: {message_content}"
            self._append_to_daily_file(message_entry, chat_name)
            
            matched_rules = self._match_regex_rules(message_content)
            if matched_rules:
                for matched_rule in matched_rules:
                    matched_data = {
                        "time": time_str,
                        "sender": sender,
                        "full_match": matched_rule["full_match"],
                        "pattern": matched_rule["pattern"],
                        "extracted_content": matched_rule["extracted_content"],
                        "extracted_contents": matched_rule.get("extracted_contents", [])
                    }
                    self._save_matched_content(matched_data, chat_name)
                    
                    logger.info("Pattern matched: %s, Extracted content: %s", matched_rule['pattern'], matched_rule['extracted_content'])
            
            logger.debug("Message saved to daily file: %s", message_entry[:100])

        except (ValueError, TypeError, AttributeError, KeyError) as e:
            logger.error("Error processing message: %s", e)
        except OSError as e:
            logger.error("OS error processing message: %s", e)
        except RuntimeError as e:
            logger.error("Runtime error processing message: %s", e)
    
    def _get_chat_info(self, chat: Any) -> Dict[str, str]:
        """Get chat information
        
        Args:
            chat: Chat object
            
        Returns:
            Dict[str, str]: Dictionary containing chat type and name
        """
        try:
            if chat and hasattr(chat, 'ChatInfo'):
                info = chat.ChatInfo()
                return {
                    "type": info.get("chat_type", ""),
                    "name": info.get("chat_name", "")
                }
        except (ValueError, TypeError, AttributeError, KeyError) as e:
            logger.error("Failed to get chat info: %s", e)
        except OSError as e:
            logger.error("OS error getting chat info: %s", e)
        except Exception as e:
            logger.error("Unexpected error getting chat info: %s", e)
        
        return {"type": "", "name": ""}

    def run(self) -> None:
        self.state.set_running(True)
        self.state.set_start_time(time.time())
        logger.info("AI worker thread started")

        try:
            if not self.init_listeners():
                self._cleanup()
                self.state.set_running(False)
                logger.error("AI worker thread initialization failed")
                return
        except (ValueError, TypeError, AttributeError, KeyError) as e:
            logger.error("Initialization failed: %s", str(e))
            self._cleanup()
            self.state.set_running(False)
            return
        except OSError as e:
            logger.error("OS error during initialization: %s", str(e))
            self._cleanup()
            self.state.set_running(False)
            return
        except RuntimeError as e:
            logger.error("Runtime error during initialization: %s", str(e))
            self._cleanup()
            self.state.set_running(False)
            return

        while not self._should_stop():
            try:
                if self.state.is_paused():
                    self.wait_for_resume()
                    continue

                if int(time.time()) % 10 == 0:  # Check every 10 seconds
                    self.update_rules()

                messages_dict = self.config.wx_instance.GetListenMessage()
                if not messages_dict:
                    time.sleep(0.1)  # Avoid high CPU usage
                    continue

                for chat, messages in messages_dict.items():
                    if self._should_stop():
                        break
                    for message in messages:
                        if self._should_stop():
                            break

                        with self.state.get_message_lock():
                            if self._is_ignored_message(message):
                                continue
                            self._process_message(message, chat)

            except (ValueError, TypeError, AttributeError, KeyError) as e:
                logger.error("Error in main loop: %s", e)
                if self._should_stop():
                    break
                time.sleep(1)  # Pause for a while when error occurs
            except OSError as e:
                logger.error("OS error in main loop: %s", e)
                if self._should_stop():
                    break
                time.sleep(1)  # Pause for a while when error occurs
            except RuntimeError as e:
                logger.error("Runtime error in main loop: %s", e)
                if self._should_stop():
                    break
                time.sleep(1)  # Pause for a while when error occurs

        self._cleanup()
        self.state.set_running(False)
        logger.info("AI worker thread stopped")

    def _is_ignored_message(self, message: Any) -> bool:
        try:
            msg_info = MessageInfo(message)
            
            if msg_info.is_system_message():
                return True

            if msg_info.is_self_message():
                return True

            if not msg_info.is_valid_message_type():
                return True

            if msg_info.is_empty_content():
                return True

            return False
        except (ValueError, TypeError, AttributeError, KeyError) as e:
            logger.error("Error checking ignored message: %s", e)
            return True
        except OSError as e:
            logger.error("OS error checking ignored message: %s", e)
            return True
        except Exception as e:
            logger.error("Unexpected error checking ignored message: %s", e)
            return True

    def _should_stop(self) -> bool:
        return self.state.get_stop_event().is_set() or not self.state.is_running()

    def _cleanup(self) -> None:
        try:
            for target in self.state.listen_list:
                if hasattr(self.config.wx_instance, "RemoveListenChat"):
                    try:
                        self.config.wx_instance.RemoveListenChat(who=target)
                    except (ValueError, TypeError, AttributeError, KeyError) as e:
                        logger.error("Failed to remove listener for %s: %s", target, e)
                    except OSError as e:
                        logger.error("OS error removing listener for %s: %s", target, e)
                    except RuntimeError as e:
                        logger.error("Runtime error removing listener for %s: %s", target, e)
            self.state.listen_list.clear()
        except (ValueError, TypeError, AttributeError, KeyError) as e:
            logger.error("Error cleaning up listeners: %s", str(e))
        except OSError as e:
            logger.error("OS error cleaning up listeners: %s", str(e))
        except RuntimeError as e:
            logger.error("Runtime error cleaning up listeners: %s", str(e))

    def pause(self) -> None:
        with self.state.get_pause_condition():
            self.state.set_paused(True)

    def resume(self) -> None:
        with self.state.get_pause_condition():
            self.state.set_paused(False)
            self.state.get_pause_condition().notify()

    def is_paused(self) -> bool:
        return self.state.is_paused()

    def wait_for_resume(self) -> None:
        with self.state.get_pause_condition():
            while self.state.is_paused():
                self.state.get_pause_condition().wait()

    def stop(self) -> None:
        self.state.get_stop_event().set()
        self.state.set_running(False)
        self.resume()

    def is_running(self) -> bool:
        return self.state.is_running()

    def get_uptime(self) -> int:
        return int(time.time() - self.state.get_start_time()) if self.state.get_start_time() else 0


class GroupWorkerManager:
    _instance: Optional['GroupWorkerManager'] = None
    workers: Dict[str, GroupWorkerThread] = {}
    lock: threading.Lock = threading.Lock()

    def __new__(cls) -> 'GroupWorkerManager':
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance.workers = {}
            cls._instance.lock = threading.Lock()
        return cls._instance

    def start_worker(self, config: WorkerConfig) -> bool:
        with self.lock:
            worker_key = f"{config.receiver}_group"
            if worker_key in self.workers:
                return False

            worker = GroupWorkerThread(config)
            self.workers[worker_key] = worker

            try:
                thread = threading.Thread(target=worker.run, daemon=True)
                thread.start()

                time.sleep(0.1)
                if not worker.is_running():
                    del self.workers[worker_key]
                    return False

                max_wait_time = 5
                wait_interval = 0.1
                total_wait_time = 0
                
                while total_wait_time < max_wait_time:
                    if not worker.is_running():
                        del self.workers[worker_key]
                        return False
                    
                    if hasattr(worker.state, 'listen_list') and len(worker.state.listen_list) > 0:
                        return True
                    
                    time.sleep(wait_interval)
                    total_wait_time += wait_interval
                
                if worker.is_running() and hasattr(worker.state, 'listen_list') and len(worker.state.listen_list) > 0:
                    return True
                else:
                    worker.stop()
                    del self.workers[worker_key]
                    return False

            except (ValueError, TypeError, AttributeError, KeyError) as e:
                if worker_key in self.workers:
                    del self.workers[worker_key]
                logger.error("Failed to start worker: %s", e)
                return False
            except OSError as e:
                if worker_key in self.workers:
                    del self.workers[worker_key]
                logger.error("OS error starting worker: %s", e)
                return False
            except RuntimeError as e:
                if worker_key in self.workers:
                    del self.workers[worker_key]
                logger.error("Runtime error starting worker: %s", e)
                return False

    def stop_worker(self, receiver: str) -> bool:
        with self.lock:
            worker_key = f"{receiver}_group"
            if worker_key in self.workers:
                worker = self.workers[worker_key]
                worker.stop()
                del self.workers[worker_key]
                
                try:
                    if os.path.exists(group_manage_config_file):
                        with open(group_manage_config_file, 'r', encoding='utf-8') as f:
                            config = json.load(f)
                        
                        config["group"] = self.config.receiver

                        with open(group_manage_config_file, 'w', encoding='utf-8') as f:
                            json.dump(config, f, ensure_ascii=False, indent=2)
                except Exception as e:
                    logger.error("Failed to update config status: %s", e)
                
                return True
            return False

    def get_worker_status(self, receiver: str) -> Optional[Dict[str, Any]]:
        with self.lock:
            worker_key = f"{receiver}_group"
            if worker_key in self.workers:
                worker = self.workers[worker_key]
                return {
                    "running": worker.is_running(),
                    "uptime": worker.get_uptime(),
                    "paused": worker.is_paused(),
                }
            return None

    def get_all_workers(self) -> List[str]:
        with self.lock:
            return list(self.workers.keys())

    def stop_all_workers(self) -> None:
        with self.lock:
            receivers = list(self.workers.keys())
            for receiver in receivers:
                self.stop_worker(receiver)

    def update_all_workers_rules(self) -> bool:
        updated_count = 0
        for _, worker in self.workers.items():
            if worker.update_rules():
                updated_count += 1

        if updated_count > 0:
            logger.info("Notified %s worker threads to update rules", updated_count)
        else:
            logger.debug("No worker threads need to update rules")

        return updated_count > 0

group_manager = GroupWorkerManager()

default_wx_instance = None

def set_wechat_instance(wx_instance):
    global default_wx_instance
    default_wx_instance = wx_instance

def get_wechat_instance():
    return default_wx_instance

def select_group(group_name: str) -> tuple:
    try:
        success, available_groups = get_available_groups()
        if not success:
            return False, "Failed to get available groups"
        
        if group_name not in available_groups:
            return False, f"Group '{group_name}' is not in the available group list"
        
        config = {}
        if os.path.exists(group_manage_config_file):
            with open(group_manage_config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
        
        config["selected_group"] = group_name
        
        with open(group_manage_config_file, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        
        logger.info("Selected group: %s", group_name)
        return True, f"Selected group: {group_name}"
    except Exception as e:
        logger.error("Failed to select group: %s", e)
        return False, f"Failed to select group: {str(e)}"

def toggle_message_recording(group_name: str, enabled: bool) -> tuple:
    try:
        config = {}
        if os.path.exists(collection_config_file):
            with open(collection_config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
        
        if group_name not in config:
            config[group_name] = {}
        config[group_name]["recording_enabled"] = enabled
        
        with open(collection_config_file, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        
        status = "enabled" if enabled else "disabled"
        logger.info("Message recording %s for %s", status, group_name)
        
        if enabled:
            group_messages_dir = os.path.join(messages_dir, group_name)
            if not os.path.exists(group_messages_dir):
                os.makedirs(group_messages_dir)
        
        return True, f"Message recording {status} for {group_name}"
    except Exception as e:
        logger.error("Failed to toggle message recording: %s", e)
        return False, f"Operation failed: {str(e)}"

def set_collection_date(group_name: str, date: str) -> tuple:
    try:
        datetime.strptime(date, "%Y-%m-%d")
        
        config = {}
        if os.path.exists(collection_config_file):
            with open(collection_config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
        
        if group_name not in config:
            config[group_name] = {}
        config[group_name]["collection_date"] = date
        
        with open(collection_config_file, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        
        logger.info("Set collection date for %s to: %s", group_name, date)
        return True, {"date": date}
    except ValueError:
        return False, "Invalid date format, please use YYYY-MM-DD format"
    except Exception as e:
        logger.error("Failed to set collection date: %s", e)
        return False, f"Operation failed: {str(e)}"

def save_collection_template(group_name: str, template: str) -> tuple:
    try:
        config = {}
        if os.path.exists(collection_config_file):
            with open(collection_config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
        
        if group_name not in config:
            config[group_name] = {}
        config[group_name]["template"] = template
        
        with open(collection_config_file, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        
        logger.info("Saved collection template for %s", group_name)
        return True, "Template saved successfully"
    except Exception as e:
        logger.error("Failed to save collection template: %s", e)
        return False, f"Operation failed: {str(e)}"

def auto_learn_pattern(group_name: str) -> tuple:
    try:
        group_messages_dir = os.path.join(messages_dir, group_name)
        if not os.path.exists(group_messages_dir):
            return False, "No message records for this group chat, cannot learn pattern"
        
        message_files = [f for f in os.listdir(group_messages_dir) if f.endswith('.json')]
        if not message_files:
            return False, "No message records for this group chat, cannot learn pattern"
        
        message_files.sort(reverse=True)
        latest_file = message_files[0]
        
        with open(os.path.join(group_messages_dir, latest_file), 'r', encoding='utf-8') as f:
            messages = json.load(f)
        
        content_patterns = []
        for msg in messages:
            content = msg.get("content", "")
            if content and len(content) > 10:
                if "=" in content and len(content.split("=")) > 1:
                    content_patterns.append("\\w+=")
                if ":" in content and len(content.split(":")) > 1:
                    content_patterns.append("\\w+:")
                
        if content_patterns:
            unique_patterns = list(set(content_patterns))
            regex = "|".join(unique_patterns)
            
            save_collection_template(group_name, regex)
            return True, f"Automatically learned pattern, generated regex: {regex}"
        else:
            return False, "Could not learn valid patterns from messages"
            
    except Exception as e:
        logger.error("Failed to auto-learn pattern: %s", e)
        return False, f"Operation failed: {str(e)}"

def get_collected_data(group_name: str, start_date: str = None, end_date: str = None) -> tuple:
    try:
        collect_dir = os.path.join(CHAT_DATE_DIR, "collect")
        
        if not os.path.exists(collect_dir):
            return False, "No chat data directory"
        
        collected_data = []
        
        for file_name in os.listdir(collect_dir):
            if not file_name.endswith('.csv'):
                continue
                
            if group_name and not file_name.startswith(group_name):
                continue
                
            if group_name:
                if file_name.startswith(group_name + "_"):
                    date_part = file_name[len(group_name + "_"):-4]
                    file_date_str = date_part
                else:
                    import re
                    date_match = re.search(r'(\d{4}-\d{2}-\d{2})', file_name)
                    if date_match:
                        file_date_str = date_match.group(1)
                    else:
                        continue
            else:
                import re
                date_match = re.search(r'(\d{4}-\d{2}-\d{2})', file_name)
                if date_match:
                    file_date_str = date_match.group(1)
                else:
                    continue
            
            if start_date is not None and file_date_str is not None and file_date_str < start_date:
                continue
            if end_date is not None and file_date_str is not None and file_date_str > end_date:
                continue
            
            file_path = os.path.join(collect_dir, file_name)
            
            with open(file_path, 'r', encoding='utf-8') as f:
                csv_reader = csv.reader(f)
                for row in csv_reader:
                    if len(row) >= 4:
                        processed_row = {
                            'time': row[0] if len(row) > 0 else '',
                            'sender': row[1] if len(row) > 1 else '',
                            'group': row[2] if len(row) > 2 else '',
                            'content': row[3] if len(row) > 3 else '',
                            'type': 'text',
                            'extractedContent': row[5] if len(row) > 5 else '',
                            'extracted_content': row[5] if len(row) > 5 else ''
                        }
                        
                        collected_data.append(processed_row)
        
        if not collected_data:
            date_range = ""
            if start_date and end_date:
                date_range = f" (from {start_date} to {end_date})"
            elif start_date:
                date_range = f" (from {start_date})"
            elif end_date:
                date_range = f" (to {end_date})"
                
            return False, f"No chat records found for {group_name}{date_range}"
        
        logger.info("Retrieved chat data for %s, total %d records", group_name, len(collected_data))
        return True, collected_data
    except Exception as e:
        logger.error("Failed to get chat data: %s", e)
        return False, f"Operation failed: {str(e)}"

def start_group_management(group_name: str, settings: dict) -> tuple:
    try:
        config = {}
        if os.path.exists(group_manage_config_file):
            with open(group_manage_config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
        
        if group_name not in config:
            config[group_name] = {}
        
        config[group_name].update({
            "enabled": True,
            "start_time": datetime.now().isoformat(),
            "settings": settings
        })
        
        config["management_enabled"] = True
        config["data_collection_enabled"] = settings.get("data_collection_enabled", False)
        config["sentiment_monitoring_enabled"] = settings.get("sentiment_monitoring_enabled", False)
        
        with open(group_manage_config_file, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        
        logger.info("Saved group management configuration for %s to group_manage.json", group_name)
        
        if default_wx_instance:
            worker_config = WorkerConfig(
                wx_instance=default_wx_instance,
                receiver=group_name,
                record_interval=settings.get("record_interval", 0),
                sensitive_words=settings.get("sensitive_words", [])
            )
            
            success = group_manager.start_worker(worker_config)
            if not success:
                return False, "Failed to start group chat monitoring"
            
            logger.info("Started monitoring group chat: %s", group_name)
        
        return True, f"Successfully started managing group chat: {group_name}"
    except Exception as e:
        logger.error("Failed to start group management: %s", e)
        return False, f"Operation failed: {str(e)}"

def start_sentiment_monitoring(group_name: str, sensitive_words: str) -> tuple:
    try:
        config = {}
        if os.path.exists(monitoring_config_file):
            with open(monitoring_config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
        
        if group_name not in config:
            config[group_name] = {}
        
        sensitive_words_list = [word.strip() for word in sensitive_words.split(',') if word.strip()]
        
        config[group_name].update({
            "enabled": True,
            "sensitive_words": sensitive_words_list,
            "start_time": datetime.now().isoformat()
        })
        
        with open(monitoring_config_file, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        
        logger.info(f"已开始{group_name}的舆情监控，敏感词: {', '.join(sensitive_words_list)}")
        return True, f"已开始舆情监控，共设置{len(sensitive_words_list)}个敏感词"
    except Exception as e:
        logger.error(f"开始舆情监控失败: {str(e)}")
        return False, f"操作失败: {str(e)}"

def stop_sentiment_monitoring(group_name: str) -> tuple:
    try:
        if not os.path.exists(monitoring_config_file):
            return False, "没有舆情监控配置"
        
        with open(monitoring_config_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        if group_name not in config:
            return False, "该群聊未开启舆情监控"
        
        config[group_name]["enabled"] = False
        config[group_name]["stop_time"] = datetime.now().isoformat()
        
        with open(monitoring_config_file, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        
        logger.info(f"已停止{group_name}的舆情监控")
        return True, "已停止舆情监控"
    except Exception as e:
        logger.error(f"停止舆情监控失败: {str(e)}")
        return False, f"操作失败: {str(e)}"

def check_sentiment_monitoring_status(group_name: str) -> tuple:
    try:
        if not os.path.exists(monitoring_config_file):
            return True, {"enabled": False}
        
        with open(monitoring_config_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        if group_name not in config or not config[group_name].get("enabled", False):
            return True, {"enabled": False}
        
        return True, {
            "enabled": True,
            "sensitive_words": config[group_name].get("sensitive_words", []),
            "start_time": config[group_name].get("start_time", ""),
            "monitoring_duration": "监控中"
        }
    except Exception as e:
        logger.error(f"检查舆情监控状态失败: {str(e)}")
        return False, f"操作失败: {str(e)}"


def get_available_groups() -> tuple:
    try:
        groups = []
        
        collect_dir = os.path.join(os.path.dirname(__file__), "chat_date", "collect")
        if os.path.exists(collect_dir):
            excel_files = [f for f in os.listdir(collect_dir) if f.endswith('.csv')]
            for file_name in excel_files:
                if '_' in file_name:
                    base_name = file_name.replace('.csv', '')
                    parts = base_name.split('_')
                    if len(parts) >= 2:
                        group_name = '_'.join(parts[:-1])
                        date_part = parts[-1]
                        
                        try:
                            datetime.strptime(date_part, "%Y-%m-%d")
                            if group_name not in groups:
                                groups.append(group_name)
                        except ValueError:
                            if base_name not in groups:
                                groups.append(base_name)
        return True, groups
        
    except Exception as e:
        logger.error(f"获取可用群组失败: {str(e)}")
        return False, f"获取群组失败: {str(e)}"


def get_group_dates(group_name: str) -> tuple:
    try:
        dates = []
        
        collect_dir = os.path.join(os.path.dirname(__file__), "chat_date", "collect")
        if os.path.exists(collect_dir):
            excel_files = [f for f in os.listdir(collect_dir) if f.endswith('.xlsx')]
            
            for file_name in excel_files:
                if '_' in file_name and group_name in file_name:
                    base_name = file_name.replace('.xlsx', '')
                    parts = base_name.split('_')
                    if len(parts) >= 2:
                        file_group_name = '_'.join(parts[:-1])
                        date_part = parts[-1]
                        
                        if file_group_name == group_name:
                            try:
                                datetime.strptime(date_part, "%Y-%m-%d")
                                if date_part not in dates:
                                    dates.append(date_part)
                            except ValueError:
                                continue
        
        dates.sort(reverse=True)
        
        return True, dates
        
    except Exception as e:
        logger.error(f"获取群组 '{group_name}' 的日期列表失败: {str(e)}")
        return False, f"获取日期列表失败: {str(e)}"


def export_collected_data_to_xlsx(group_name: str, start_date: str = None, end_date: str = None) -> tuple:
    try:
        import re
        import json
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "收集数据"
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        headers = ["时间", "发送者", "群聊", "消息内容", "消息类型"]
        extracted_headers = []
        
        success, data = get_collected_data(group_name, start_date, end_date)
        if not success:
            return False, data
            
        max_extracted_items = 0
        for item in data:
            extracted_content = item.get('extractedContent', '') or item.get('extracted_content', '')
            if extracted_content:
                items = []
                if isinstance(extracted_content, str) and extracted_content.startswith('[') and extracted_content.endswith(']'):
                    try:
                        items = json.loads(extracted_content)
                        if not isinstance(items, list):
                            items = [items]
                    except json.JSONDecodeError:
                        items = [item.strip() for item in extracted_content.split('，') if item.strip()]
                else:
                    items = [item.strip() for item in extracted_content.split('，') if item.strip()]
                
                max_extracted_items = max(max_extracted_items, len(items))
        
        for i in range(max_extracted_items):
            extracted_headers.append(f"提取内容{i+1}")
        
        all_headers = headers + extracted_headers
        for col_num, header in enumerate(all_headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for row_num, item in enumerate(data, 2):
            ws.cell(row=row_num, column=1, value=item.get('time', ''))
            ws.cell(row=row_num, column=2, value=item.get('sender', ''))
            ws.cell(row=row_num, column=3, value=item.get('group', ''))
            ws.cell(row=row_num, column=4, value=item.get('content', ''))
            ws.cell(row=row_num, column=5, value=item.get('type', ''))
            
            extracted_content = item.get('extractedContent', '') or item.get('extracted_content', '')
            if extracted_content:
                items = []
                if isinstance(extracted_content, str) and extracted_content.startswith('[') and extracted_content.endswith(']'):
                    try:
                        items = json.loads(extracted_content)
                        if not isinstance(items, list):
                            items = [items]
                    except json.JSONDecodeError:
                        items = [item.strip() for item in extracted_content.split('，') if item.strip()]
                else:
                    items = [item.strip() for item in extracted_content.split('，') if item.strip()]
                    
                for col_idx, extracted_item in enumerate(items, 6):
                    ws.cell(row=row_num, column=col_idx, value=str(extracted_item))
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        export_dir = os.path.join(os.path.dirname(__file__), "chat_date", "export")
        if not os.path.exists(export_dir):
            os.makedirs(export_dir)
            
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if group_name:
            file_name = f"{group_name}_导出数据_{timestamp}.xlsx"
        else:
            file_name = f"全部群聊_导出数据_{timestamp}.xlsx"
            
        file_path = os.path.join(export_dir, file_name)
        wb.save(file_path)
        
        logger.info(f"数据已导出为xlsx文件: {file_path}")
        return True, file_path
        
    except Exception as e:
        logger.error(f"导出数据为xlsx失败: {str(e)}")
        return False, f"导出失败: {str(e)}"


regex_config_file = os.path.join(data_dir, "regex_rules.json")
